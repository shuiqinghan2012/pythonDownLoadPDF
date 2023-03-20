from openpyxl import load_workbook
from mylog import Mylog,logging
import os
log = Mylog()
log.configLog()

"""
处理Excel文件数据的描述
1.读取excel 活动sheet的所有数据

"""

logging.info("开始FileExcel程序")
class FileExcel():
    """
    一个描述Excel二维表的处理，将excel数据写入二维列表并返回列表
    """
    def __init__(self,excelPath):
        self.path = excelPath
        # print("path=", self.path)
        self.workBook = load_workbook(self.path)
        self.sheet = self.workBook.active  # workBook["sheet1"]
        self.maxRow = self.sheet.max_row
        self.maxCol = self.sheet.max_column
        #按表格形式存储整个表格内容的二维列表
        self.fileList = []
        #按列表存储PartPN信息
        self.partList = []

    """
    读取填写查找文件名的Excel文件
    返回Excel文件中（活动表格）所有的数据，按二维列表返回
    """
    def readExcelToList(self):
        #返回Excel文件中所有的文件的列表
        maxRow = self.maxRow
        maxCol = self.maxCol


        for rowNum in range(1,maxRow + 1):#按行访问Excel
            # print('row is ',rowNum)
            #存储行数据的列表
            rowList = []
            for colNum in range(1,maxCol + 1):
                # print('col is ',colNum)
                # print('row,col is ',rowNum,colNum)
                # print(self.sheet.cell(rowNum,colNum).value)#打印cell的值
                rcValue = self.sheet.cell(rowNum, colNum).value
                #数据非空则加入行列表
                if rcValue is not None:
                    rowList.append(rcValue)
                    # print(rowList)
            #将完整的一行数据列表加入到整个表格的二维列表中
            self.fileList.append(rowList)
        #返回整个二维列表
        return self.fileList

    @staticmethod
    def getWordXYInList(wordList,word):
        # logging.info('start getWordXYInList({}{})\n'.format(word, wordList))
        for countRow in range(0, len(wordList)):
            # list中查找特定字符串，返回下标 1218修改
            for countCol in range(0, len(wordList[countRow])):
                if (word == wordList[countRow][countCol]):
                    print(word, countRow, countCol)
                    logging.info('getWordXYInList 查找{}所在的list中行和列为({},{})\n'.format(word, countRow, countCol))
                    print(wordList[countRow][countCol])
                    return countRow, countCol
                    break
                else:
                    continue
        # if( countRow = len(wordList)-1 && countCol = len(wordList[countRow])-1)
        #             logging.debug('在{}中未找到{}'.format(wordList,word))
        #             #print('在{}中未找到{}'.format(wordList,word))
        #             return None
    def getColWord(self,PN):
        logging.info('从列表中获取需要的全部列数据!partPN')
        partRow, partCol = self.getWordXYInList(self.fileList,PN)
        # print(self.fileList[int(partRow) +1:])
        for rowList in self.fileList[int(partRow) +1:]:
            if len(rowList) > partCol:
                # print(rowList[partCol])
                self.partList.append(rowList[partCol])
        # print(self.partList)
        return self.partList
    #将数据data写入到excel sheet中的row，col单元格中
    def writeExcelCell(self, data,row=None, col=None):
        logging.info("writeExcelCell")
        # print("writeExcelCell",data, row, col)
        if (row and col):  # 都不是None
            self.sheet.cell(row, col,data)
            # print("在{}{}行列写入{}".format(row, col,data))
            return True
        if ((row == None) and (col == None)):  # 都是None
            print("((row == None) and (col == None))")
            return False
        # 如下分别特定行写入和特定列入的方式还不确定如何做，data应为list
        # elif (row != None):
        #     for col in range(1, maxCol + 1):
        #         # self.sheet.cell(row, col,data)
        #
        # elif (col != None):
        #     for row in range(1, maxRow + 1):
        #         # self.sheet.cell(row, col,data)

    def closeExcel(self):
        self.workBook.close()

    def saveResult(self,name=None):
        self.workBook.close()
        if name:
            t_name = "copy_"+name
            print(t_name)
        self.workBook.save(t_name)



# **********************************测试Demo************************
# excelPath =os.path.join(os.getcwd(),'SAPexport20210324.XLSX')
#
# #测试readExcelToList ok
# xmj = FileExcel(excelPath)
# fileList = xmj.readExcelToList()
# print('fileList:',fileList)
# for file in fileList:
#     print(file)
