"""调用Lxexcel相关类的函数及上层函数
处理文件夹移动，文件移动，共有四种情况
1.文件夹名包含在路径中
2.文件夹名和路径最后的文件夹名完全相同
3.文件名包含在文件路径中
4.文件名和去除后缀的文件名完全相同

"""
import os,datetime
import shutil
from openpyxl import load_workbook
from mylog import Mylog,logging
# import logging, sys, traceback
from FileExcel import FileExcel

import re

righnote ='''
Copyright [2022] by [Michael]
1999-2025 All rights reserved.
'''
log = Mylog()
log.configLog()
logging.info("开始主程序")
#*********************************路径处理************************************
print(righnote)
debug= False
# print(os.getcwd())
excelPath = os.getcwd()
print(excelPath)
logging.info("excelPath ：{}".format(excelPath))

# path = os.getcwd()




#******************************路径处理****************************
class LibExcel():
    def __init__(self, excelPath):
        self.path = excelPath
        # print("path=", self.path)
        self.workBook = load_workbook(self.path)
        self.sheet = self.workBook.active  # workBook["sheet1"]
        self.maxRow = self.sheet.max_row
        self.maxCol = self.sheet.max_column
        logging.info('sheet,max_row,max_column {},{},{}'.format(self.sheet,self.sheet.max_row,self.sheet.max_column))
    '''
    关闭excel
    '''
    def closeExcel(self):
        self.workBook.save(excelName)
        self.workBook.close()
    '''getCellColum函数获得strValue文本所对应的列
       返回 列坐标数字
    '''
    def getCellColum(self, strValue):
        logging.info('start getCellColum {}'.format(strValue))
        # 查找位号所在的列
        for rowNum in range(1, self.sheet.max_row):
            for cell in self.sheet[rowNum]:
                if (cell.value) == strValue:
                    if (debug):
                        # print(cell.coordinate)
                        print(cell.column)
                        # print(cell.value)
                    return cell.column
    #修改电阻value的内容
    def modifyValue(self,value):
        # 去掉空格
        value = value.replace(" ", "")
        #mΩ电阻使用mR标识
        if 'mΩ' in value:
            value = value.replace("mΩ","mR")
        # 去掉Ω
        value = value.replace("Ω", "")
        # 去掉±
        value = value.replace("±", ";")
        # 去掉SMD
        value = value.replace("SMD", ";")
        print(value)
        return value
    #向Excel特定cell写入值
    def writeValue(self,row,col,value):
        # 写入对应的value 列
        logging.info('向Excel中写入value {}'.format(value))
        self.sheet.cell(row, col, value)


    '''setResValue 通过电阻描述找到对应的符合规则描述的value，
        输入：标题行所在的关键词 getKeyValue= Description
    '''
    def setResValue(self,strValue):
        column = self.getCellColum(strValue)
        valueCol = self.getCellColum('Value')
        print(valueCol)
        #遍历打印每行的描述
        logging.info('excel sheet is  {}'.format(self.sheet))
        for row in range(2,self.sheet.max_row + 1):
            descrip = self.sheet.cell(row,column).value
            print(descrip)
            #处理电阻描述的正则表达
            valuePattern = re.compile(r'(\d+([.]\d+)?)\w?[Ω]\s?[±](\d+([.]\d+)?)[%]')
            packagePattern = re.compile(r'SMD\s\d+')
            valueMo = valuePattern.search(descrip)
            packageMo =packagePattern.search(descrip)
            if valueMo is not None:
                #如匹配成功处理完后写入对应的excel 列
                value = valueMo.group() + packageMo.group()
                value = self.modifyValue(value)
                self.writeValue(row,valueCol,value)

    def packageModify(self,package):
        # 去掉空格
        package = package.replace(" ", "")
        # 标准格式为SR0201
        package = package.replace("SMD", "SR")
        return package

    '''setResPackage 通过电阻描述找到对应的符合规则描述的封装值，
        输入：标题行所在的关键词 getKeyValue= Description
    '''
    def setResPackage(self,strValue):
        column = self.getCellColum(strValue)
        packageCol = self.getCellColum('Allegro PCB Footprint')
        print("packageCol is %d",(packageCol))
        for row in range(2,self.sheet.max_row + 1):
            descrip = self.sheet.cell(row,column).value
            # print(descrip)
            # 处理电阻封装的正则表达
            packagePattern = re.compile(r'SMD\s\d+')
            packageMo = packagePattern.search(descrip)
            if packageMo is not None:
                #如果匹配成功处理完后写入对应的excel 列
                packge = packageMo.group()
                packge = self.packageModify(packge)
                self.writeValue(row, packageCol, packge)


    def findcap(self):
        #查找cap类型，并将其按主要参数分类返回
        for line in self.fileList:
            if ("Chip Capacitor" or "片状陶瓷电容") in line[1]:
                #将第二个元素分解成新列表后插入原列表切片的位置
                lineSplit = line[2].split(";")
                line[2:2]=lineSplit
                # print(line)
                logging.info(line)
                self.result.writeList(line)

        self.result.saveResult()


#*********************start****************************

'''
处理电阻value和footprint的程序
value = 86.6K;1%;0201
footpr = SR0201
Description = 厚膜电阻 86.6KΩ ±1% 1/20W SMD 0201(公制0603)

输入配置为
'''
#输入配置
excelName = '副本电容_Capacitors_EDADOC.xlsx'
excelPathFile =os.path.join(os.getcwd(),excelName)

# 处理电阻value和footprint
libRes = LibExcel(excelPathFile)
# desColum = libCap.getCellColum('Description')
# print(desColum)
libRes.setResValue('Description')
libRes.setResPackage('Description')
#关闭excel
libRes.closeExcel()










# func = getattr(findFile,fun)
# func()

# try:
#     run()
# except Exception as err:
#     logging.exception(err)
